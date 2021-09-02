# -*- coding: utf-8 -*-
"""
Created on Fri Jul 12 09:50:00 2019

This module contains all the functions used in the modules IV.py, CV.py, IVmultiple.py, CVmultiple.py, Vbd.py. 

@author: Sneha Ramshanker 
"""

"""
Importing packages
"""
import Constants as con
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import csv
from scipy import stats
from scipy.optimize import fsolve
import seaborn as sns
from scipy.interpolate import UnivariateSpline
from openpyxl import load_workbook
from sklearn.linear_model import LinearRegression

"""
Functions
"""
def breakdownVol(df, x, y, thresh = 0.9995):
    """
    Parameters:
        df: Pandas dataframe 
        x: String - name of x column in df (voltage)
        y: String - name of y column in df (capacitance)
    Optional Parameters: 
        thresh: Float - r squared threshold value
    Returns:
        breakdownVol: Float - Breakdown Voltage 
    Note: 
        This function determines the breakdown by calculation the point where the relationship between the first and second derivative becomes extremely linear. 
    """
    frst_der = manual_der(df[x].values, df[y].values)
    scnd_der = manual_der(df[x].values, frst_der)
    
    r_value = 1
    N = 2
    while round(r_value,4) >= thresh: #Threshold for linearity
        X = np.log(frst_der[-N:])
        Y = np.log(scnd_der[-N:])
        slope, intercept, r_value, p_value, std_err = stats.linregress(X,Y)
        N = N+1
          
    voltage = df[x].values
    breakdownVol = voltage[-N+1]
    return breakdownVol
        
def calculate_width(C, A):
    """
    Parameters:
        C: Float - aapacitance 
        A: Float - area of the detector 
    
    Returns: 
        integer - width of capacitor
    """
    
    #C = eps * A / d
    return (0 if (C == 0) else ((con.eps_si * con.eps0 * A) / C))

def cleanupCV(df, x, y):
    """
    Parameters:
        df: Pandas dataframe 
        x: String - name of x column in df (voltage)
        y: String - name of y column in df (capacitance)
    Returns:
        df: Pandas dataframe 
    Note: 
        The data is cleaned up so that it can be stored nicely in a dataframe. The cleanup proceduce is specific to the format of the textfile the CV data is stored in. 
    """
    df = df.iloc[0:len(df)-1]

    df[x] = pd.to_numeric(df[x])
    df[y] = pd.to_numeric(df[y])

    df[x] = abs(df[x])
    df[y] = abs(df[y])

    dfclean = (df.dropna(subset = [x, y]))
    return df 

def cleanupIV(df, x, y):
    """
    Parameters:
        df: Pandas dataframe 
        x: String - name of x column in df (voltage)
        y: String - name of y column in df (capacitance)
    Returns:
        dfclean: Pandas dataframe 
    Note: 
        The data is cleaned up so that it can be stored nicely in a dataframe. The cleanup proceduce is specific to the format of the textfile the IV data is stored in. 
    """
    df = df.iloc[0:len(df)-1]

    df[x] = pd.to_numeric(df[x])
    df.loc[y] = pd.to_numeric(df.loc[:,y])

    # This complains, but I'm not sure what to do about it.
    df[x] = abs(df[x])
    df[y] = abs(df[y])

    df["Outlier"] = ""
    dfclean = (df.dropna(subset = [x, y]))
    return dfclean  

def dataplot(df, x, y,image_path, log, xlabel, ylabel, iden = "None", gaindeplvol = 0, totdeplvol = 0, gainwidth = 0, breakdownVol = 0):
    """
    Parameters: 
        df: Pandas Dataframe 
        x: String - name of x column in df 
        y: String - name of y column in df 
        image_path: String - path where the image will be saved
        log: String ("log" or "ylog" or "nolog") - should the x axis or y axis be log scale? "log" makes xaxis log scaled and "ylog" makes yaxis log scaled.
        xlabel: String - label of xaxis 
        ylabel: String - label of yaxis 
    Optional Parameters:
        iden: String - identification tag for plotting additional elements  ["DPWcut", "invCV", "IV"]
        gaindeplvol: Float - value of depletion voltage of gain layer 
        totdeplvol: Float - value of total depletion voltage
        gainwidth: Float - value of peak of gain layer (measured in microns)
        breakdownVol: Float - value of breakdown voltage
    Returns:
        No returns but saves the plot     
    Note: 
        The following function produces and saves a x-y plot from  data in a dataframe . 
        Given additional optional parameters, the function can produce more complex plots. 
    """
    plt.figure(figsize=(12,8)) 
    plt.plot(df[x][1:], df[y][1:], '-')
    if log == "log":
        plt.xscale('log')
    if log == "ylog":
        plt.yscale('log')
    if iden == "DPWcut":
        plt.xlim(0,5)
    if iden == "invCV":
        plt.axvline(gaindeplvol, linestyle = '--', color = 'forestgreen', label = 'Gain Depl Vol: '+ str(round(gaindeplvol,1)))
        #plt.axvline(totdeplvol, linestyle = '--', color = 'darkorchid', label = 'Total Depl Vol: '+ str(round(totdeplvol,1)))
    if iden == "IV":
        plt.axvline(breakdownVol, linestyle = '--', color = 'k', alpha = 0.5, label = 'Breakdown V: '+ str(round(breakdownVol,1)) + " V")
    plt.legend()
    plt.xlabel(xlabel)
    plt.ylabel(ylabel)
    plt.savefig(image_path)
    plt.clf()
    plt.close()

def dataplotDPW(df, x, y,image_path, log, xlabel, ylabel, gainwidth, gainpeak,  FWHM = [0, 0, 0]):
    """
    Parameters:
        df: Pandas dataframe
        x: String - name of x column in df 
        y: String - name of y column in df 
        image_path: String - path where the image will be saved
        xlabel: String - label of xaxis 
        ylabel: String - label of yaxis 
        gainwidth: Float - width value of peak of gain layer (measured in microns)
        gainpeak: Float - Doping profile value of the peak of gain layer (measured in (cm^-3) )
    Optional Parameters:
        FWHM: list [fwhm, x, g] where:
            fwhm: float - full width half maximum 
            x: numpy array -  x values of plot 
            g: numpy array - gaussian fit generated using astropy 
    Returns:
        No returns but saves plot 
    Note:
        This function is specific for plotting Doping Profile vs Width 
    """
    plt.figure(figsize=(12,8)) 
    plt.plot(df[x], df[y], '-')
    if log == "log":
        plt.xscale('log')
    if log == "ylog":
        plt.yscale('log')
    plt.xlabel(xlabel)
    plt.ylabel(ylabel)
    plt.vlines(df[x].max(), df[y].min(), df[y].max() + 0.2*(df[y].max()), 'darkorchid', linestyles = "dashed", label = "Total Depth: "+ str(round(df[x].max(),2)) + r"$ \mu m $")
    plt.vlines(gainwidth, df[y].min(), df[y].max() + 0.2*(df[y].max()), 'forestgreen', linestyles = "dashed", label = "Peak Depth: "+ str(round(gainwidth,2)) + r"$\mu m$")
    plt.hlines(gainpeak, 0,  gainwidth, 'deepskyblue', linestyles = "dashed", label = "Peak Value: "+ str('%.2g' % gainpeak) + r"$ cm^{-3}$")

    if FWHM != [0, 0, 0]:
        plt.plot([], [], ' ', label="Gain width (FWHM): "+ str(round(FWHM[0], 2))+ r"$ \mu m $")        
        plt.plot(FWHM[1], FWHM[2](FWHM[1]),'--y', label='Gaussian Fit')
    plt.legend(loc = 'upper center')
    
    plt.savefig(image_path)
    plt.clf()
    plt.close()

def dataplotIV(df,xd,yd,hued, impathIV, breakdownVol):
    """
    Parameters:
        df: Pandas Dataframe 
        xd: String - name of x column in df 
        yd: String - name of y column in df 
        hued: String - name of hue column in df (Seaborn functionality for lmplot)
        impathIV: String - path where image will be saved 
        breakdownVol: Float - Breakdown Voltage 
    Return:
        No returns but image is saved 
    Note:
        This is an outdated function that plots a visualiztion of how breakdown voltage is determined by plotting two linear fits of the data points and determining the point of intersection.
    """
    g = sns.lmplot(x = xd, y = yd, hue = hued, data = df, aspect = 1.5, legend = False )
    plt.axvline(breakdownVol, linestyle = '--', color = 'k', label = 'Breakdown Voltage: '+ str(round(breakdownVol,1)))
    #g.set( yscale="log")
    #g.subplots(figsize=(20,15))
    g.set(ylim=(0, 5*10**-9))
    plt.legend(fontsize = 'large')
    g.set_axis_labels('Voltage (V)', 'Current (A)')
    g.savefig(impathIV)

def FWHM(arr_x,arr_y):
    """
    Parameters:
        arr_x: Numpy Array - x array of gaussian fit 
        arr_y: Numpy Array - y array of gaussian fit 
    Return:
        FWHM: float - full width half maximum 
    Note:
        This function returns the full width half maximum of a Gaussian Fit 

    """
    difference = max(arr_y) - min(arr_y)
    HM = difference / 2
    
    #pos_extremum = arr_y.argmax()  # or in your case: arr_y.argmin()
    pos_extremum = arr_y.argmin()  # or in your case: arr_y.argmin()
    #print difference, HM, pos_extremum, arr_y, arr_x, arr_y[pos_extremum:-1]
    
    nearest_above = (np.abs(arr_y[pos_extremum:-1] - HM)).argmin()
    nearest_below = (np.abs(arr_y[0:pos_extremum] - HM)).argmin()
    #print difference, HM, pos_extremum, nearest_above, nearest_below
    
    FWHM = (np.mean(arr_x[nearest_above + pos_extremum]) - 
            np.mean(arr_x[nearest_below]))
    return FWHM

def findIntersection(s1,in1,s2, in2, estimate = 0):
    """
    Parameters:
        s1: Float - slope of first line
        s2: Float - slope of second line
        in1: Float - y intercept of first line
        in2: Float - y intercept of second line 
    Optional Parameters
        estimate: Float - estimate of the point of intersection (improves accuracy of the result)
    Return
        float - x value of the point of intersection
    """
    return fsolve(lambda x : (s1*x+in1) - (s2*x+in2),estimate)

def gen_der(spline): 
    """
    Parameters:
        spline: UnivariateSpline
    Return:
        UnivariateSpline
    Note:
        The function calculates the derivative of an inputted function
    """
    return spline.derivative()

def gen_spline(x,y):
    """
    Parameters:
        x: list or array 
        y: list or array 
    Return:
        spline
    Note: 
        function produces a spline given x and y arrays
    """   
    return UnivariateSpline(x, y, k=3, s=0)

def get_doping_profile(CV,area, voltageName, capacitanceName):
    """
    Parameters:
        CV: Pandas Dataframe - 2 column dataframe with column names {'voltage', 'capacitance'}
        area: Float 
    Return:
        df: Pandas Dataframe
    Note: 
        The following function used the CV curve to obtain the doping profile
    """
    
    width, profile = [], []
    for n in range(2,CV.shape[0]):
        slope = lin_reg(CV.iloc[n-1:n+1,:],voltageName,capacitanceName)[0]*10**12
        C = CV.iloc[n,1]*10**12
        constant = con.eps_si*con.q*con.eps0*area**2
        if slope == 0:
            N = 0 
        else:
            N = abs((1/slope)*(C**3/(constant)))*10**-24  
        w = calculate_width(CV[capacitanceName][n], area)*10**4
        width.append(w) #Reporting in um
        profile.append(N) #Reporting in cm^-3            
    df = pd.DataFrame(columns=['width', 'profile'])
    df['width'] = width
    df['profile'] = profile
    return df

def isoutlier(df, a, p, column, start = 0):
    """
    Parameters:
        df: Pandas Dataframe 
        a: Integer - number of standard deviations to determine outlier 
        p: Float - fraction of data to be used to determine mean and standard deviation 
        column: String - name of the column for the current 
    Optional Parameters:
        start: Integer - index the scanning for outliers begins at 
    Return:
        df: Pandas Dataframe
    Note: 
        This function is used to fit two lines to IV curves by classifying each point as outliers or not i.e. in the Baseline region or the Breakdown region
    """
    #a sets the number of standard deviations to determine outlier 
    #N is the number of points over which mean and stdev should be determined
    #Picking a small subset of datapoints 
    #N = round(p * df.shape[0])
    
    df_sub = df.sort_values(column).loc[strt_idx:end_idx] 
    df = df.loc[turn_on_curve:]
    mean = df_sub[column].mean()
    std = df_sub[column].std()
    imin = df.index[0]
    imax = df.index[-1]   
    #Determining if the points are outliers 
    for i in range (imin, imax):
        x = df.loc[i,column] 
        #((mean-a*std) <= x)
        if (mean-a*std) <= x and x<= (mean + a*std):
            df.loc[i, "Outlier"] = 'Baseline Region'
        else:
            df.loc[i,"Outlier"] = 'Breakdown Region'
    df['Outlier'].replace('', np.nan, inplace=True)
    df['Outlier'].dropna()
    return df

def isoutlierCV (df, a, p, column, direction):
    """
    Parameters:
        df: Pandas dataframe 
        a: Integer - number of standard deviations to determine outlier 
        p: Float - fraction of data to be used to determine mean and standard deviation 
        column: String - name of the column for the current 
        direction: {'lr', 'rl'} - Direction to scan for outliers (left to right or right to left)
    Return:
        df: Pandas Dataframe 
    Note:
        This function is used to classify points and outliers or not in order to fit two lines to the data and find the point of intersection. 
    """
    # Trying to find where the linear fit starts breaking down
    # Basically looking for when the prediction doesn't fit the lowest point that is being fit by a lot more than it doesn't fit the other points
    for i in range(len(df[column].values)-3, 0, -1):
      slope, intercept, r_value, p_value, std_err = stats.linregress(df[0].values[i:len(df[column].values)], df[column].values[i:len(df[column].values)])
      predictions = intercept + slope*df[0].values
      mymean =  (abs((df[column].values[i:len(df[column].values)]-predictions[i:len(df[column].values)] )/ df[column].values[i:len(df[column].values)])).mean()
      if (abs((df[column].values[i:len(df[column].values)]-predictions[i:len(df[column].values)] )/ df[column].values[i:len(df[column].values)]))[0] > 5*mymean:
         break
       
    print i, df[0].values[i]
    ibreakdown = i+1

    df_sub = df
    mean = df_sub[column].mean()
    std = df_sub[column].std()
    imin = df.index[0]
    imax = df.index[-1]
    df["Outlier"] = "n"
    
    for i in range(ibreakdown, len(df[column].values)):
        df.loc[i, "Outlier"] = 'y'
    for i in range(0, ibreakdown-3):
        df.loc[i, "Outlier"] = 'm'
   
    return df

def lin_reg(df,x,y):
    """
    Parameters: 
        df: Pandas Dataframe 
        x: String - name of x column in df 
        y: String - name of y column in df
    Return:
        [slope, intercept]: List 
        slope: float 
        intercept: float 
    Note:
        This function performs linear regression on a dataframe 
    """
    Y = df[y].values
    X = df[x].values
    slope, intercept, r_value, p_value, std_err = stats.linregress(X,Y)
    return [slope, intercept]

def manual_der (X, Y):
    """
    Parameters:
        X: List/Array
        Y: List/Array 
    Return
        der: List/Array - derivative dY/dX 
    Note:
        This function performs manual differentiation to calculate dY/dX
    """
    der = []
    for n in range(0, len(X)):
        #slope, intercept, r_value, p_value, std_err = stats.linregress(X[n:n+1],Y[n:n+1])
        
        if n == 0:
            slope, intercept, r_value, p_value, std_err = stats.linregress(X[n:n+2],Y[n:n+2])
        else:
            slope, intercept, r_value, p_value, std_err = stats.linregress(X[n-1:n+1],Y[n-1:n+1])
        der.append(slope)
        
    return der 

def nextopen(value, column, sheet):    
    """
    Parameters:
        value: String/Integer/Float - Value to be inputted in the excel sheet 
        column: String - Column name on excel sheet. Example : 'A'
        sheet: String - Sheet name on excel. Example: "Sheet1"
    Returns:
        None
    Note:
        This function inputs a given value in the next open cell in a specified column of an excel sheet 
    """
    cnum = ord(column)-64
    if sheet.cell(row=sheet.max_row, column= cnum).value == None:        
        sheet[column + str(sheet.max_row)] = value
    else:
        sheet[column + str(sheet.max_row+1)] = value

def reg_intersect (df1, df2, x, y, estimate = 0 ):
    """
    Parameters:
        df1: Pandas Dataframe 
        df2: Pandas Dataframe 
        x: String - name of x column for df1 and df2 (should be the same) 
        y: String - name of y column for df1 and df2 (should be the same)
    Optional Parameters:
        estimate: estimated value for the point of intersection 
    Return: 
        float - x value of the point of intersection
    Note:
        This function determines the point of intersection of two lines described by two dataframes
    """
    #print df1, x, y
    [s1,i1] = lin_reg(df1,x, y)
    [s2,i2] = lin_reg(df2,x, y)
    #return [s1, i1, s2, i2]
    return findIntersection(s1, i1, s2, i2, estimate)

def scatterplot(df, x, y):
    """
    Parameters:
        df: Pandas Dataframe 
        x: String - name of x column in df 
        y: String - name of y column in df
    Returns:
        None
    Note:
        The function plots a scatter plot given a dataframe and the x and y column names 
    """
    plt.plot(df[x],df[y],'o')
    plt.xlabel(x)
    plt.ylabel(y)

def simpletextread(file, delim): #FOR DEBUGGING
    """
    Parameters:
        file: String - path of file 
        delim: String - delimiter used to separate columns. Example '\t'
    Return:
        data: Pandas Dataframe 
    Note:
        This function converts a file with a specified delimiter to a pandas dataframe. It is really useful for debugging
    """
    data = pd.read_csv(file, delimiter = delim)
    return data

def splitdata(df):
    """
    Parameters: 
        df: Pandas Dataframe 
    Return:
        Concatenated Pandas Dataframe 
        return[0] = df_out: Pandas Dataframe 
        return[1] = df_nout: Pandas Dataframe 
    Note:
        This function splits a dataframe into two depending on whether it is a outlier or not {Outlier: 'y', Not Outlier: 'n'}
    """
    df_out = df.loc[df['Outlier'] =='y']
    df_nout = df.loc[df['Outlier'] =='n']
    return(df_out, df_nout)
    
def storedataCV(fileName, uniqueID):
    """
    Parameters:
        df: Pandas Dataframe 
    Return:
        df: Pandas Dataframe 
    Note:  
        This function stores CV data into a dataframe
    """
    # Find the row that corresponds to the unique ID
    count = 0
    with open(fileName) as myfile:
      for line in myfile:
        count += 1
        if line.startswith(uniqueID):
           break

    df = pd.read_csv(fileName, skiprows = count, sep="\t", header = None) #Store data into dataframe 
    #df = pd.read_csv(fileName, skiprows = count, sep=",", header = None) #Store data into dataframe 
    return df 

def storedataIV(fileName, uniqueID):
    """
    Parameters:
        df: Pandas Dataframe 
    Return:
        df: Pandas Dataframe 
    Note:  
        This function stores IV data into a dataframe
    """

    # Find the row that corresponds to the unique ID
    count = 0
    with open(fileName) as myfile:
      for line in myfile:
        count += 1
        if line.startswith(uniqueID):
           break

    df = pd.read_csv(fileName, skiprows = count, sep="\t", header = None) #Store data into dataframe 
    return df 

def setPlotStyle():
    """
    Parameters:
        None
    Return: 
        None
    Note:
        This function sets are the styllistic parameters for plots 
    """
    plt.style.use('ggplot')
    plt.rcParams['lines.linewidth'] = 2.15
    plt.rcParams['lines.markeredgewidth'] = 0.0
    plt.rcParams['font.size'] = 14
    plt.rcParams['axes.labelsize'] = 18
    plt.rcParams['axes.labelweight'] = 'bold'
    plt.rcParams['axes.facecolor'] = 'white'
    plt.rcParams['grid.color'] = 'black'
    plt.rcParams['grid.linestyle'] = '--'
    plt.rcParams['grid.linewidth'] = '0.25'
    plt.rcParams['grid.alpha'] = '0.25'
    plt.rcParams['xtick.labelsize'] = 16
    plt.rcParams['ytick.labelsize'] = 16
    plt.rcParams['legend.fontsize'] = 12
    plt.rcParams['legend.frameon'] = False
    plt.rcParams['figure.titlesize'] = 'large'
    plt.rcParams['figure.titleweight'] = 'bold'
    plt.rcParams['figure.figsize'] = (10, 10)
    plt.rcParams['axes.edgecolor'] = 'black'
    plt.rcParams['patch.edgecolor'] = 'none'
    plt.rcParams.update({'figure.max_open_warning': 0})
    plt.rcParams['legend.fontsize'] = 20


def tablebreakdownVol(dest,breakdownVol, column, sheet):
    """
    Parameters:
        dest: String - path of excel file 
        breakdownVol: Float - Breakdown voltage value 
        column: String - Name of excel column to store data 
        sheet: String - Name of excel sheet to store data 
    Return:
        None
    Note:
        This function stores breakdown voltage in specified location in an excel sheet 
    """
    #Open an xlsx for reading
    wb = load_workbook(filename = dest)
    #Setting working sheet
    ws = wb.get_sheet_by_name(sheet)   
    #Storing in next open cell in the column
    nextopen(breakdownVol,column,ws)
    #ws['B' + str(3)] = 2
    wb.save(dest)

#Outdated Functions 
def calcbreakdownVol(df,x, y): 
    """
    Parameters:
        df: Pandas dataframe 
        x: String - name of x column in df (Voltage)
        y: String - name of y column in df (Current)

    Returns:
        breakdownVol: integer - Breakdown Voltage 

    Note: 
        This function does not calculate breakdown voltage well
    """
    x = df[x].values
    y = df[y].values
    frst_der = gen_der(gen_spline(x,y))
    scnd_der = gen_der(gen_der(gen_spline(x,y)))
    thrd_der = gen_der(scnd_der)
    
    df2 = pd.DataFrame({'x':x, 'y':thrd_der(x)})
    dupli = df2[df2['y'].duplicated() == True]

    i = -1 
    while dupli['y'].iloc[i] == dupli['y'].iloc[i-1]:
        index = (df2.shape[0]-1)+(i-1)
        breakdownVol = (df2['x'].iloc[index])
        i = i-1
        if i < -10: 
            break
    return breakdownVol


        
